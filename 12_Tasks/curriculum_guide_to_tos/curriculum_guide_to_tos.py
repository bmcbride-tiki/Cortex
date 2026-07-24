"""
---
module_id: 'TRANSFORM-CG-TOS'
title: 'Curriculum Guide to TOS Workbook Transformer'
tags: ['automation', 'parser', 'docx', 'xlsx']
schema_version: '1.0.0'
---
Self-contained, UI-free engine for converting Curriculum Guide Word documents into TOS Excel workbooks.
"""

from __future__ import annotations

import getpass
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl.cell.cell import Cell, MergedCell

PERIOD_WORD_TO_NUMBER = {
    "ONE": 1,
    "TWO": 2,
    "THREE": 3,
    "FOUR": 4,
    "FIVE": 5,
    "SIX": 6,
}


@dataclass
class TopicData:
    """One topic inside a Curriculum Guide section."""

    topic_id: str
    topic_letter: str
    topic_name: str
    hours: Optional[float]
    percentage: Optional[float]


@dataclass
class SectionData:
    """One section inside a Curriculum Guide period."""

    section_title: str
    section_hours: Optional[float]
    section_percentage: Optional[float]
    topics: List[TopicData]


@dataclass
class PeriodData:
    """One period of training from the Curriculum Guide."""

    period_name: str
    period_number: Optional[int]
    period_hours: Optional[float]
    sections: List[SectionData]


@dataclass
class CurriculumGuideData:
    """Parsed Curriculum Guide data."""

    trade_name: str
    periods: List[PeriodData]


@dataclass
class ConversionSettings:
    """Settings for one conversion run."""

    curriculum_guide_file: Path
    template_file: Path
    output_folder: Path
    trade_name_override: str
    exam_questions: int


@dataclass
class RowCapacityIssue:
    """A section/core competence block that does not have enough topic rows."""

    sheet_name: str
    period_name: str
    section_number: int
    section_title: str
    available_rows: int
    required_rows: int

    @property
    def rows_to_add(self) -> int:
        """Return the minimum number of rows that should be added."""

        return max(0, self.required_rows - self.available_rows)


class CurriculumGuideParser:
    """
    Parser for Alberta-style Curriculum Guide Word documents.

    The parser looks for:
    - PERIOD ONE COURSE CONTENT
    - SECTION ONE:
    - Topic A.
    - hours and percentages in nearby table text
    """

    def __init__(self) -> None:
        self.period_pattern = re.compile(
            r"^PERIOD\s+(ONE|TWO|THREE|FOUR|FIVE|SIX)\s+COURSE\s+CONTENT$",
            re.IGNORECASE,
        )
        self.period_hours_pattern = re.compile(
            r"\(?\d+\s*weeks?\s*[–-]\s*(\d+(?:\.\d+)?)\s*hours?\)?",
            re.IGNORECASE,
        )
        self.section_pattern = re.compile(
            r"^SECTION\s+(ONE|TWO|THREE|FOUR|FIVE|SIX|XXX):",
            re.IGNORECASE,
        )
        self.topic_pattern = re.compile(
            r"^Topic\s+([A-Z\d]+)\.",
            re.IGNORECASE,
        )
        self.hours_pattern = re.compile(
            r"(\d+(?:\.\d+)?)\s*(?:Hours?|hrs)",
            re.IGNORECASE,
        )
        self.percent_pattern = re.compile(
            r"(\d+(?:\.\d+)?)\s*%",
            re.IGNORECASE,
        )

    def parse(self, file_path: Path, trade_name_override: str = "") -> CurriculumGuideData:
        """Parse a Curriculum Guide .docx file."""

        document = Document(str(file_path))
        trade_name = trade_name_override.strip() or self._guess_trade_name(
            document=document,
            file_path=file_path,
        )

        periods: List[PeriodData] = []
        current_period: Optional[PeriodData] = None
        current_section: Optional[SectionData] = None
        period_section_weights: Dict[str, Dict[str, Optional[float]]] = {}
        look_for_period_hours = 0

        for block in self._iter_block_items(document):
            if isinstance(block, Paragraph):
                text = self._clean_text(block.text)

                if not text:
                    continue

                period_match = self.period_pattern.match(text)

                if period_match:
                    period_word = period_match.group(1).upper()

                    current_period = PeriodData(
                        period_name=text,
                        period_number=PERIOD_WORD_TO_NUMBER.get(period_word),
                        period_hours=None,
                        sections=[],
                    )
                    periods.append(current_period)

                    current_section = None
                    period_section_weights = {}
                    look_for_period_hours = 4
                    continue

                if current_period is not None and look_for_period_hours > 0:
                    hours_match = self.period_hours_pattern.search(text)

                    if hours_match:
                        current_period.period_hours = self._clean_number(hours_match.group(1))
                        look_for_period_hours = 0
                    else:
                        look_for_period_hours -= 1

                section_match = self.section_pattern.match(text)

                if section_match and current_period is not None:
                    section_word = section_match.group(1).upper()
                    weights = period_section_weights.get(section_word, {"hours": None, "percentage": None})

                    current_section = SectionData(
                        section_title=text,
                        section_hours=self._as_optional_float(weights.get("hours")),
                        section_percentage=self._as_optional_percentage(weights.get("percentage")),
                        topics=[],
                    )
                    current_period.sections.append(current_section)

            elif isinstance(block, Table):
                if current_period is not None and not current_period.sections:
                    self._extract_summary_weights(table=block, weights_dict=period_section_weights)

                if current_section is not None:
                    topic = self._extract_topic_from_table(block)

                    if topic is not None:
                        current_section.topics.append(topic)

        cleaned_periods = [
            period for period in periods if period.sections or period.period_hours is not None
        ]

        return CurriculumGuideData(trade_name=trade_name, periods=cleaned_periods)

    def _guess_trade_name(self, document: Any, file_path: Path) -> str:
        """Guess the trade name from the first useful paragraphs."""

        for paragraph in document.paragraphs[:20]:
            text = self._clean_text(paragraph.text)

            if not text:
                continue

            lowered = text.lower()

            if any(
                skip_text in lowered
                for skip_text in ["apprenticeship", "curriculum guide", "course outline", "table of contents"]
            ):
                continue

            cleaned = text.replace("Description of the ", "")
            cleaned = cleaned.replace(" Trade", "")
            cleaned = cleaned.strip()

            if cleaned:
                return cleaned

        return file_path.stem

    def _iter_block_items(self, parent: Any) -> Any:
        """
        Yield paragraphs and tables in document order.

        python-docx normally separates paragraphs and tables. This helper keeps
        the original document order so the parser knows which topics belong to
        which section.
        """

        from docx.oxml.table import CT_Tbl
        from docx.oxml.text.paragraph import CT_P

        parent_element = parent.element.body

        for child in parent_element.iterchildren():
            if isinstance(child, CT_P):
                yield Paragraph(child, parent)
            elif isinstance(child, CT_Tbl):
                yield Table(child, parent)

    def _extract_summary_weights(self, table: Table, weights_dict: Dict[str, Dict[str, Optional[float]]]) -> None:
        """Extract section hours and percentage from a period summary table."""

        try:
            for row in table.rows:
                first_cell_text = self._clean_text(row.cells[0].text).upper()
                match = re.search(r"SECTION\s+(ONE|TWO|THREE|FOUR|FIVE|SIX)", first_cell_text)

                if not match:
                    continue

                section_word = match.group(1)
                hours_found: Optional[float] = None
                percentage_found: Optional[float] = None

                for cell in row.cells[1:]:
                    value = self._clean_text(cell.text)

                    percent_match = self.percent_pattern.search(value)

                    if percent_match:
                        percentage_found = self._percent_text_to_decimal(percent_match.group(1))
                        continue

                    number_match = re.search(r"(\d+(?:\.\d+)?)", value)

                    if number_match:
                        hours_found = self._clean_number(number_match.group(1))

                weights_dict[section_word] = {"hours": hours_found, "percentage": percentage_found}

        except Exception:
            return

    def _extract_topic_from_table(self, table: Table) -> Optional[TopicData]:
        """Extract one topic from a topic table."""

        try:
            first_cell_text = self._clean_text(table.cell(0, 0).text)
            topic_match = self.topic_pattern.match(first_cell_text)

            if not topic_match:
                return None

            topic_letter = topic_match.group(1).upper()
            topic_id = first_cell_text
            topic_name = ""

            if len(table.rows[0].cells) > 1:
                topic_name = self._clean_text(table.cell(0, 1).text)

            topic_hours: Optional[float] = None
            topic_percentage: Optional[float] = None

            for row in table.rows:
                row_text = " ".join(
                    self._clean_text(cell.text) for cell in row.cells if self._clean_text(cell.text)
                )

                hours_match = self.hours_pattern.search(row_text)
                percent_match = self.percent_pattern.search(row_text)

                if hours_match:
                    topic_hours = self._clean_number(hours_match.group(1))

                if percent_match:
                    topic_percentage = self._percent_text_to_decimal(percent_match.group(1))

            return TopicData(
                topic_id=topic_id,
                topic_letter=topic_letter,
                topic_name=topic_name,
                hours=topic_hours,
                percentage=topic_percentage,
            )

        except Exception:
            return None

    def _clean_number(self, value: Any) -> Optional[float]:
        """Convert a value into a float when possible."""

        if value is None:
            return None

        match = re.search(r"(\d+(?:\.\d+)?)", str(value).replace(",", ""))

        if not match:
            return None

        try:
            return float(match.group(1))
        except ValueError:
            return None

    def _percent_text_to_decimal(self, value: Any) -> Optional[float]:
        """Convert percentage number text to a decimal."""

        number = self._clean_number(value)

        if number is None:
            return None

        return number / 100

    def _as_optional_float(self, value: Any) -> Optional[float]:
        """Return value as float or None."""

        if value is None or value == "":
            return None

        if isinstance(value, (int, float)):
            return float(value)

        return self._clean_number(value)

    def _as_optional_percentage(self, value: Any) -> Optional[float]:
        """Return a percentage value as a decimal."""

        if value is None or value == "":
            return None

        if isinstance(value, (int, float)):
            numeric_value = float(value)

            if numeric_value > 1:
                return numeric_value / 100

            return numeric_value

        text = str(value)

        if "%" in text:
            return self._percent_text_to_decimal(text)

        number = self._clean_number(text)

        if number is None:
            return None

        if number > 1:
            return number / 100

        return number

    def _clean_text(self, text: Any) -> str:
        """Normalize whitespace."""

        if text is None:
            return ""

        return re.sub(r"\s+", " ", str(text)).strip()


class TOSWorkbookWriter:
    """
    Writes Curriculum Guide data into the TOS Excel template.

    The writer looks for existing TOS blocks that begin with a Section or Core
    Competence row and end at a Subtotals row. The May 2026 TOS template uses
    rows such as "Section 1" and "Subtotals Section 1", so block detection is
    based on those markers rather than the older core-competence-only markers.

    If a section has more topics than the template has rows for, the extra
    topics beyond the template's capacity are skipped and a warning is logged
    rather than stopping the conversion.
    """

    def write(self, parsed_data: CurriculumGuideData, settings: ConversionSettings, log=print) -> Path:
        """Create the output TOS workbook."""

        settings.output_folder.mkdir(parents=True, exist_ok=True)

        output_path = self._build_output_path(settings)
        shutil.copy2(settings.template_file, output_path)

        workbook = openpyxl.load_workbook(output_path)

        try:
            issues = self.check_row_capacity(workbook=workbook, parsed_data=parsed_data)

            if issues:
                log(
                    "Template row capacity warning: some sections have more topics than available "
                    "template rows. The conversion will continue and extra topics beyond template "
                    "capacity will be skipped."
                )
                for issue in issues:
                    log(
                        f"  {issue.sheet_name} / {issue.period_name} / Section {issue.section_number}: "
                        f"available rows {issue.available_rows}, required {issue.required_rows}. "
                        f"Skipped topics beyond the available template rows: {issue.rows_to_add}."
                    )

            for period in parsed_data.periods:
                sheet_name = self._find_sheet_for_period(workbook=workbook, period=period)

                if sheet_name is None:
                    log(f"WARNING: No matching sheet found for {period.period_name}.")
                    continue

                worksheet = workbook[sheet_name]

                self._write_period_sheet(
                    worksheet=worksheet,
                    trade_name=parsed_data.trade_name,
                    period=period,
                    settings=settings,
                )

            self._add_conversion_summary(workbook=workbook, parsed_data=parsed_data, settings=settings)

            workbook.save(output_path)

        except Exception:
            workbook.close()

            if output_path.exists():
                try:
                    output_path.unlink()
                except Exception:
                    pass

            raise

        workbook.close()

        return output_path

    def check_row_capacity(self, workbook: Any, parsed_data: CurriculumGuideData) -> List[RowCapacityIssue]:
        """Check whether each TOS section block has enough rows."""

        issues: List[RowCapacityIssue] = []

        for period in parsed_data.periods:
            sheet_name = self._find_sheet_for_period(workbook=workbook, period=period)

            if sheet_name is None:
                continue

            worksheet = workbook[sheet_name]
            section_blocks = self._find_section_blocks(worksheet)

            for section_index, section in enumerate(period.sections):
                if section_index >= len(section_blocks):
                    issues.append(
                        RowCapacityIssue(
                            sheet_name=sheet_name,
                            period_name=period.period_name,
                            section_number=section_index + 1,
                            section_title=section.section_title,
                            available_rows=0,
                            required_rows=len(section.topics),
                        )
                    )
                    continue

                block = section_blocks[section_index]
                available_rows = self._count_available_topic_rows(block)
                required_rows = len(section.topics)

                if required_rows > available_rows:
                    issues.append(
                        RowCapacityIssue(
                            sheet_name=sheet_name,
                            period_name=period.period_name,
                            section_number=section_index + 1,
                            section_title=section.section_title,
                            available_rows=available_rows,
                            required_rows=required_rows,
                        )
                    )

        return issues

    def _count_available_topic_rows(self, block: Dict[str, int]) -> int:
        """Count the available supporting-competency/topic rows in one block."""

        first_topic_row = block["first_topic_row"]
        last_topic_row = block["last_topic_row"]

        if last_topic_row < first_topic_row:
            return 0

        return last_topic_row - first_topic_row + 1

    def _find_sheet_for_period(self, workbook: Any, period: PeriodData) -> Optional[str]:
        """Find the workbook sheet for a parsed period."""

        candidates: List[str] = []

        if period.period_number is not None:
            candidates.extend(
                [f"Period {period.period_number}", f"PERIOD {period.period_number}", str(period.period_number)]
            )

        candidates.append(period.period_name)

        normalized_sheet_lookup = {
            self._normalize_sheet_name(sheet_name): sheet_name for sheet_name in workbook.sheetnames
        }

        for candidate in candidates:
            normalized_candidate = self._normalize_sheet_name(candidate)

            if normalized_candidate in normalized_sheet_lookup:
                return normalized_sheet_lookup[normalized_candidate]

        return None

    def _normalize_sheet_name(self, value: Any) -> str:
        """Normalize sheet names for matching."""

        return re.sub(r"[^a-z0-9]+", "", str(value).lower())

    def _write_period_sheet(self, worksheet: Any, trade_name: str, period: PeriodData, settings: ConversionSettings) -> None:
        """Write one period into its matching TOS sheet."""

        self._write_period_header(worksheet=worksheet, trade_name=trade_name, period=period, settings=settings)

        section_blocks = self._find_section_blocks(worksheet)

        for section_index, block in enumerate(section_blocks):
            if section_index >= len(period.sections):
                continue

            section = period.sections[section_index]

            self._write_section_block(
                worksheet=worksheet,
                block=block,
                section=section,
                section_number=section_index + 1,
            )

    def _write_period_header(self, worksheet: Any, trade_name: str, period: PeriodData, settings: ConversionSettings) -> None:
        """Write period-level header values."""

        set_cell_value_safely(worksheet=worksheet, row_number=1, column_number=1, value=trade_name.upper())

        if period.period_hours is not None:
            set_cell_value_safely(worksheet=worksheet, row_number=3, column_number=4, value=period.period_hours)

        set_cell_value_safely(worksheet=worksheet, row_number=3, column_number=6, value=settings.exam_questions)

        self._replace_trade_placeholders(worksheet=worksheet, trade_name=trade_name)

    def _write_section_block(self, worksheet: Any, block: Dict[str, int], section: SectionData, section_number: int) -> None:
        """Write one Curriculum Guide section into one TOS section block."""

        section_row = block["section_row"]
        first_topic_row = block["first_topic_row"]
        last_topic_row = block["last_topic_row"]

        set_cell_value_safely(worksheet=worksheet, row_number=section_row, column_number=1, value=section.section_title)
        set_cell_value_safely(worksheet=worksheet, row_number=section_row, column_number=4, value=section.section_percentage)
        set_cell_value_safely(worksheet=worksheet, row_number=section_row, column_number=6, value=section.section_hours)

        available_rows = list(range(first_topic_row, last_topic_row + 1))

        for topic_index, row_number in enumerate(available_rows):
            if topic_index >= len(section.topics):
                continue

            topic = section.topics[topic_index]

            set_cell_value_safely(
                worksheet=worksheet,
                row_number=row_number,
                column_number=1,
                value=section_number if topic_index == 0 else None,
            )
            set_cell_value_safely(worksheet=worksheet, row_number=row_number, column_number=2, value=topic.topic_letter)
            set_cell_value_safely(worksheet=worksheet, row_number=row_number, column_number=3, value=topic.topic_name)
            set_cell_value_safely(worksheet=worksheet, row_number=row_number, column_number=7, value=topic.hours)
            set_cell_value_safely(worksheet=worksheet, row_number=row_number, column_number=10, value=topic.percentage)

    def _find_section_blocks(self, worksheet: Any) -> List[Dict[str, int]]:
        """
        Find section/core competence blocks in a TOS sheet.

        Supports:
            Core Competence 1 [Insert Core Competence Title]
            Section One: Foundational Skills...
        """

        section_rows: List[int] = []
        subtotal_rows: List[int] = []

        for row_number in range(1, worksheet.max_row + 1):
            value = get_cell_value_safely(worksheet=worksheet, row_number=row_number, column_number=1)

            if not isinstance(value, str):
                continue

            cleaned = value.strip().lower()

            if cleaned.startswith("core competence") or cleaned.startswith("section"):
                section_rows.append(row_number)

            if cleaned.startswith("subtotals"):
                subtotal_rows.append(row_number)

        blocks: List[Dict[str, int]] = []

        for section_row in section_rows:
            subtotal_row = None

            for possible_subtotal_row in subtotal_rows:
                if possible_subtotal_row > section_row:
                    subtotal_row = possible_subtotal_row
                    break

            if subtotal_row is None:
                continue

            blocks.append(
                {
                    "section_row": section_row,
                    "first_topic_row": section_row + 1,
                    "last_topic_row": subtotal_row - 1,
                    "subtotal_row": subtotal_row,
                }
            )

        return blocks

    def _replace_trade_placeholders(self, worksheet: Any, trade_name: str) -> None:
        """Replace visible [Trade Name] placeholders."""

        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue

                if "[Trade Name]" not in cell.value:
                    continue

                new_value = cell.value.replace("[Trade Name]", trade_name)
                set_cell_value_safely(worksheet=worksheet, row_number=cell.row, column_number=cell.column, value=new_value)

    def _add_conversion_summary(self, workbook: Any, parsed_data: CurriculumGuideData, settings: ConversionSettings) -> None:
        """Add a summary sheet for review."""

        summary_sheet_name = "CG to TOS Summary"

        if summary_sheet_name in workbook.sheetnames:
            del workbook[summary_sheet_name]

        sheet = workbook.create_sheet(summary_sheet_name)

        sheet["A1"] = "Curriculum Guide to TOS Conversion Summary"
        sheet["A3"] = "Trade Name"
        sheet["B3"] = parsed_data.trade_name
        sheet["A4"] = "Source Curriculum Guide File"
        sheet["B4"] = settings.curriculum_guide_file.name
        sheet["A5"] = "Converted At"
        sheet["B5"] = datetime.now().isoformat(timespec="seconds")
        sheet["A7"] = "Period"
        sheet["B7"] = "Section"
        sheet["C7"] = "Section Hours"
        sheet["D7"] = "Section %"
        sheet["E7"] = "Topic"
        sheet["F7"] = "Topic Hours"
        sheet["G7"] = "Topic %"

        row_number = 8

        for period in parsed_data.periods:
            for section in period.sections:
                if not section.topics:
                    sheet.cell(row_number, 1).value = period.period_name
                    sheet.cell(row_number, 2).value = section.section_title
                    sheet.cell(row_number, 3).value = section.section_hours
                    sheet.cell(row_number, 4).value = section.section_percentage
                    row_number += 1
                    continue

                for topic in section.topics:
                    sheet.cell(row_number, 1).value = period.period_name
                    sheet.cell(row_number, 2).value = section.section_title
                    sheet.cell(row_number, 3).value = section.section_hours
                    sheet.cell(row_number, 4).value = section.section_percentage
                    sheet.cell(row_number, 5).value = f"{topic.topic_letter} {topic.topic_name}".strip()
                    sheet.cell(row_number, 6).value = topic.hours
                    sheet.cell(row_number, 7).value = topic.percentage
                    row_number += 1

        for column_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            sheet.column_dimensions[column_letter].width = 28

        sheet.column_dimensions["E"].width = 70

    def _build_output_path(self, settings: ConversionSettings) -> Path:
        """Build the output filename."""

        user_name = getpass.getuser().strip() or "User"
        today = datetime.now().strftime("%m-%d-%Y")
        base_name = f"{user_name} - Curriculum Guide to TOS - {today}.xlsx"

        output_path = settings.output_folder / self._safe_filename(base_name)

        if not output_path.exists():
            return output_path

        counter = 2

        while True:
            candidate = output_path.with_name(f"{output_path.stem} ({counter}){output_path.suffix}")

            if not candidate.exists():
                return candidate

            counter += 1

    def _safe_filename(self, filename: str) -> str:
        """Make a Windows-safe filename."""

        clean = re.sub(r'[<>:"/\\|?*]', "_", filename)
        clean = re.sub(r"\s+", " ", clean).strip()
        return clean


def set_cell_value_safely(worksheet: Any, row_number: int, column_number: int, value: Any) -> None:
    """
    Safely set a worksheet cell value.

    If the target cell is part of a merged range, Excel only allows writing to
    the top-left cell of that merged range.
    """

    target_cell = worksheet.cell(row=row_number, column=column_number)

    if isinstance(target_cell, Cell):
        target_cell.value = value
        return

    merged_top_left = find_merged_range_top_left(worksheet=worksheet, row_number=row_number, column_number=column_number)

    if merged_top_left is None:
        return

    top_left_row, top_left_column = merged_top_left
    worksheet.cell(row=top_left_row, column=top_left_column).value = value


def get_cell_value_safely(worksheet: Any, row_number: int, column_number: int) -> Any:
    """Safely read a worksheet cell value."""

    target_cell = worksheet.cell(row=row_number, column=column_number)

    if not isinstance(target_cell, MergedCell):
        return target_cell.value

    merged_top_left = find_merged_range_top_left(worksheet=worksheet, row_number=row_number, column_number=column_number)

    if merged_top_left is None:
        return None

    top_left_row, top_left_column = merged_top_left
    return worksheet.cell(row=top_left_row, column=top_left_column).value


def find_merged_range_top_left(worksheet: Any, row_number: int, column_number: int) -> Optional[Tuple[int, int]]:
    """Find the top-left cell of the merged range containing a cell."""

    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row_number <= merged_range.max_row
            and merged_range.min_col <= column_number <= merged_range.max_col
        ):
            return merged_range.min_row, merged_range.min_col

    return None


CURRENT_TOOL_DIR = Path(__file__).resolve().parent
VAULT_ROOT = CURRENT_TOOL_DIR.parents[1]
OUTBOX_DIR = VAULT_ROOT / "99_Outbox" / "curriculum_guide_to_tos"


def run_conversion(
    curriculum_guide_file: Path,
    trade_name_override: str = "",
    exam_questions: int = 100,
    template_file: Optional[Path] = None,
    output_folder: Optional[Path] = None,
    log=print,
) -> Path:
    """Parse a Curriculum Guide file and write the resulting TOS workbook."""

    settings = ConversionSettings(
        curriculum_guide_file=curriculum_guide_file,
        template_file=template_file or (CURRENT_TOOL_DIR / "TOS_Template_May_2026.xlsx"),
        output_folder=output_folder or OUTBOX_DIR,
        trade_name_override=trade_name_override,
        exam_questions=exam_questions,
    )

    if not settings.curriculum_guide_file.exists():
        raise FileNotFoundError(f"Curriculum Guide file not found: {settings.curriculum_guide_file}")

    if settings.curriculum_guide_file.suffix.lower() == ".doc":
        raise ValueError(
            "Legacy .doc files are not supported. Save the document as .docx (Word: "
            "File > Save As > Word Document) and try again."
        )

    if not settings.template_file.exists():
        raise FileNotFoundError(f"TOS template not found: {settings.template_file}")

    log("Parsing Curriculum Guide Word document...")

    parser = CurriculumGuideParser()
    parsed_data = parser.parse(file_path=settings.curriculum_guide_file, trade_name_override=settings.trade_name_override)

    if not parsed_data.periods:
        raise ValueError("No periods were found in the selected Curriculum Guide file.")

    log(f"Found {len(parsed_data.periods)} period(s).")

    for period in parsed_data.periods:
        log(f"  {period.period_name}: {len(period.sections)} section(s)")

    log("Locating TOS template section rows...")

    writer = TOSWorkbookWriter()
    output_path = writer.write(parsed_data=parsed_data, settings=settings, log=log)

    log(f"Generated TOS workbook: {output_path}")

    return output_path


def _log(message: str) -> None:
    print(f"[CG-TO-TOS] {message}", flush=True)


def _prompt_curriculum_guide_file() -> Path:
    while True:
        value = input("Enter the path to the Curriculum Guide .docx file: ").strip().strip('"')

        if not value:
            print("  A file path is required. Try again.")
            continue

        path = Path(value)

        if not path.exists():
            print(f"  File not found: {path}. Try again.")
            continue

        return path


if __name__ == "__main__":
    cg_file = Path(sys.argv[1]) if len(sys.argv) > 1 else _prompt_curriculum_guide_file()
    trade_name_arg = sys.argv[2] if len(sys.argv) > 2 else ""
    exam_questions_arg = int(sys.argv[3]) if len(sys.argv) > 3 else 100

    try:
        result_path = run_conversion(
            curriculum_guide_file=cg_file,
            trade_name_override=trade_name_arg,
            exam_questions=exam_questions_arg,
            log=_log,
        )
    except Exception as err:
        _log(f"Conversion stopped: {err}")
        sys.exit(1)

    _log(f"Output written to: {result_path}")
